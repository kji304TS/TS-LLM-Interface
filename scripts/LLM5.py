import requests
import json
from datetime import datetime, timedelta
import re
import os
from dotenv import load_dotenv
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import openpyxl.utils
import argparse
import sys
import pytz
from textblob import TextBlob  
import nltk # Import nltk
import aiohttp
import asyncio
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import logging

# Set up logging configuration
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'debug_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# --- Download NLTK resources if not already present ---
# TextBlob relies on these for POS tagging and tokenization.
REQUIRED_NLTK_RESOURCES = ['punkt', 'averaged_perceptron_tagger']

try:
    # Check if resources are available by trying to find them
    for resource in REQUIRED_NLTK_RESOURCES:
        nltk.data.find(f"tokenizers/{resource}" if resource == 'punkt' else f"taggers/{resource}")
    print("✅ NLTK resources (punkt, averaged_perceptron_tagger) found.")
except nltk.downloader.DownloadError as e:
    print(f"⚠️ NLTK DownloadError ({e}). This might be an issue if resources are truly missing and download fails.")
except LookupError:
    print("ℹ️ NLTK resources not found, attempting to download...")
    try:
        for resource in REQUIRED_NLTK_RESOURCES:
            nltk.download(resource, quiet=True)
        print("✅ Successfully downloaded NLTK resources (punkt, averaged_perceptron_tagger).")
        # Re-check after download attempt
        for resource in REQUIRED_NLTK_RESOURCES:
            nltk.data.find(f"tokenizers/{resource}" if resource == 'punkt' else f"taggers/{resource}")
    except Exception as e_download:
        print(f"❌ FAILED to download NLTK resources. TextBlob functionalities (POS tagging, keyword analysis) might fail. Error: {e_download}")
        print("Please try manually downloading them: import nltk; nltk.download('punkt'); nltk.download('averaged_perceptron_tagger')")
# --- End NLTK Resource Download ---

# Dynamically add the project root to sys.path
# This allows finding the 'utils' module when the script is run from the 'scripts' directory
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from utils.slack_notifier import send_slack_report
from utils.intercom_team_fetcher import get_intercom_teams # Added import

# ✅ Load .env variables
load_dotenv()  # <-- This must be called BEFORE using os.getenv()

# ✅ Get values from .env
# API_KEY = os.getenv("API_KEY") # Handled by app.py's GDrive upload
# GDRIVE_FOLDER_ID = os.getenv("GDRIVE_FOLDER_ID") # Handled by app.py's GDrive upload
INTERCOM_PROD_KEY = os.getenv("INTERCOM_PROD_KEY") # Still needed for Intercom calls

# --- CONFIGURABLE CONSTANTS ---
# Actual Intercom custom attribute name for product area (ensure this matches Intercom)
ACTUAL_PRODUCT_AREA_CUSTOM_ATTRIBUTE_NAME = "MetaMask area"

# Actual Intercom custom attribute key for Team Inbox
TEAM_INBOX_CUSTOM_ATTRIBUTE_KEY = "Team Inbox"

# Custom attribute field name for elevated conversations
ELEVATED_BY_FIELD_NAME = "elevated_by"

# Team to Team Inbox mapping
SCRIPT_TEAM_TO_TEAM_INBOX_VALUE = {
    "MetaMask TS": "MetaMask HD Technical Support",  # Updated with actual value
    "Card": "MetaMask HD Card",
    "Portfolio": "MetaMask HD Portfolio",
    "Solana": "MetaMask HD Solana",
    "MetaMask HD UST": "MetaMask HD UST"
}

USA_PRIMARY_TIMEZONE = "America/Chicago"
# --- END CONFIGURABLE CONSTANTS ---

CATEGORY_HEADERS = {
    "Bridges": ["Bridge Issue"],
    "Card": ["MM Card Issue", "MM Card Partner issue", "Dashboard Issue", "KYC Issue", "Dashboard Issue - Subcategory", "KYC Issue - Subcategory"],
    "Dashboard": ["Dashboard issue"],
    "Ramps": ["Buy or Sell", "Buy issue", "Sell issue"],
    "SDK": [],
    "Security": ["scam vector", "Phishing Method", "Funds missing", "Compromise Method"], # Added "Compromise Method"
    "Snaps": ["Snaps Category"],
    "Staking": ["Staking Feature", "Validator Staking Issue", "Pooled Staking Issue", "Liquid Staking Issue", "Third Party Staking", "Bug ID", "Refund amount (USD)", "Refund Provided", "Withdrawals", "Managing Staked Tokens", "User Training", "Failed Transaction", "Liquid Staking Provider", "Staking Token Type", "Staking Platform"],
    "Swaps": ["Swaps issue"],
    "Wallet": ["Wallet issue"],
    "Wallet API": [],
    "Portfolio": [], 
    "Solana": []     
}

OUTPUT_DIR = "output_files"
INSIGHTS_DIR = "Outputs"
TEAM_REPORTS_DIR = os.path.join(INSIGHTS_DIR, "team_reports") # New directory for team reports

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(INSIGHTS_DIR, exist_ok=True)
os.makedirs(TEAM_REPORTS_DIR, exist_ok=True) # Create team_reports directory

# ✅ Predefined Prompts - Restructured
PREDEFINED_PROMPTS = {
    "GLOBAL": { # Prompts applicable to most/all reports
        "Trends": [
            "How many conversations occurred in each subcategory?", # This might need to be more generic or context-aware
            "What percentage of total issues does each subcategory represent?", # Ditto
            "How have issue frequencies changed over time?",
            "What correlations exist between different issue types?",
            "Are there seasonal trends in user-reported issues?"
        ],
        "Keyword Analysis": [
            "What are the top 10 most important keywords?", # Removed 'in the summaries' for generality
            "What sentiment trends can be observed?" # Removed 'from the summaries'
        ],
        "Conversation Volume": [
            "How many conversations are in each MetaMask area?", # This is specific, might move or rephrase
            "Which MetaMask area has seen the highest increase in conversations?" # Ditto
        ]
    },
    "Security": { # Prompts specific to the Security product area
        "Scam Vector Analysis": [
            "What is the most frequent value in the 'scam vector' field?",
            "What are the counts for each 'scam vector' value (e.g., Address Poisoning, Airdrop Claim, etc.)?",
            "What percentage of Security conversations does each 'scam vector' represent?",
            "Are there any notable keywords or summaries associated with high-frequency 'scam vector' types?"
        ],
        "Phishing Method Analysis": [ # New category for Phishing Method
            "What is the most frequent value in the 'Phishing Method' field?",
            "What are the counts for each 'Phishing Method' value (e.g., Angler phishing, Email phishing, etc.)?",
            "What percentage of Security conversations does each 'Phishing Method' represent?",
            "Are there any notable keywords or summaries associated with high-frequency 'Phishing Method' types?"
        ],
        "Funds Missing Analysis": [
            "What is the most frequent reason in the 'Funds missing' field (e.g., SRP/PKcompromised, User error, Report scam, No funds lost)?",
            "What are the counts for each 'Funds missing' reason (e.g., SRP/PKcompromised, Unintended contract interaction, User error, Ecosystem exploit, No information, No funds lost, Unknown, Blockaid false positive, eth-phishing-detect false positive, Law enforcement / legal inquiry, Report scam, Request for update on prior ticket, Security questions)?",
            "What percentage of Security conversations does each 'Funds missing' reason represent?",
            "Are there notable keywords or summaries associated with high-frequency 'Funds missing' reasons, especially for actual loss categories vs. 'No funds lost' categories?"
        ],
        "Compromise Method Analysis": [ # New category for SRP/PK Compromise Methods
            "If 'Funds missing' is 'SRP/PKcompromised', what is the most frequent value in the 'Compromise Method' field (e.g., Malware, Rotten Seed)?",
            "What are the counts for each 'Compromise Method' (e.g., Malware, Rotten Seed, SRP Digitally Stolen, SRP Phished Directly, SRP Physically Stolen)?",
            "What percentage of SRP/PK compromised conversations does each 'Compromise Method' represent?",
            "Are there specific keywords or summary patterns associated with different 'Compromise Method' types like 'Malware' or 'Rotten Seed'?"
        ],
        "Top Issues": [ # Security might still have other general issue columns, if applicable
            # Add prompts here if Security uses other columns from CATEGORY_HEADERS like "Security Issue Subcategory"
        ]
    },
    "Card": {
        "Top Issues": [
            "What is the most frequent subcategory in the 'MM Card Issue' column?",
            "What is the most frequent subcategory in the 'MM Card Partner issue' column?",
            "What is the most frequent subcategory in the 'Dashboard Issue' column (related to Card)?",
            "What is the most frequent subcategory in the 'KYC Issue' column (related to Card)?"
        ]
    },
    "Bridges": {
        "Top Issues": ["What is the most frequent subcategory in the 'Bridge Issue' column?"]
    },
    "Ramps": {
        "Top Issues": [
            "What is the most frequent subcategory in the 'Buy issue' column?",
            "What is the most frequent subcategory in the 'Sell issue' column?"
        ]
    },
    "Snaps": {
        "Top Issues": ["What is the most frequent subcategory in the 'Snaps Category' column?"]
    },
    "Staking": {
        "Top Issues": [
            "What is the most frequent subcategory in the 'Staking Feature' column?",
            "What is the most frequent subcategory in the 'Validator Staking Issue' column?",
            "What is the most frequent subcategory in the 'Pooled Staking Issue' column?",
            "What is the most frequent subcategory in the 'Liquid Staking Issue' column?",
            "What is the most frequent subcategory in the 'Third Party Staking' column?"
        ]
    },
    "Swaps": {
        "Top Issues": ["What is the most frequent subcategory in the 'Swaps issue' column?"]
    },
    "Wallet": {
        "Top Issues": ["What is the most frequent subcategory in the 'Wallet issue' column?"]
    }
    # Add other product areas (Dashboard, SDK, Wallet API, Portfolio, Solana) if they have specific columns/prompts
}

# Removed get_last_week_dates() - dates will be passed from app.py
# Helper function to format dates for file naming
def get_yyyymmdd_date_strings(date_str_from_app):
    # Assumes date_str_from_app is "YYYY-MM-DD HH:MM"
    dt_obj = datetime.strptime(date_str_from_app.split(" ")[0], "%Y-%m-%d")
    return dt_obj.strftime("%Y%m%d")

# ✅ Extract and clean text
def remove_html_tags(text):
    if not isinstance(text, str):
        return ''
    clean = re.sub(r'<.*?>', '', text)
    return clean

def sanitize_text(text):
    if text:
        return text.replace('\u200b', '').encode('utf-8', 'ignore').decode('utf-8')
    return text

# ✅ Fetch summaries and transcripts
def get_conversation_summary(conversation):
    if 'conversation_parts' in conversation:
        conversation_parts = conversation['conversation_parts'].get('conversation_parts', [])
        for part in conversation_parts:
            if part.get('part_type') == 'conversation_summary':
                return remove_html_tags(part.get('body', ''))
    return "No summary available"

def get_conversation_transcript(conversation):
    transcript = []
    if 'conversation_parts' in conversation:
        conversation_parts = conversation['conversation_parts'].get('conversation_parts', [])
        for part in conversation_parts:
            if part.get('part_type') == 'comment':
                author = part.get('author', {}).get('type', 'Unknown')
                comment = remove_html_tags(part.get('body', ''))
                transcript.append(f"{author}: {comment}")
    return "\n".join(transcript) if transcript else "No transcript available"

# --- Session Management ---
def create_session():
    """Creates a requests session with retry logic and connection pooling."""
    session = requests.Session()
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504]
    )
    adapter = HTTPAdapter(
        max_retries=retry_strategy,
        pool_connections=10,
        pool_maxsize=10
    )
    session.mount("https://", adapter)
    return session

# Global session for reuse
SESSION = create_session()

# --- Cache for frequently accessed data ---
@lru_cache(maxsize=1000)
def get_cached_conversation(conversation_id):
    """Cached version of get_intercom_conversation."""
    return get_intercom_conversation(conversation_id)

async def fetch_conversation_page(session, url, headers, payload, page_num):
    """Fetch a single page of conversations asynchronously."""
    try:
        async with session.post(url, headers=headers, json=payload) as response:
            if response.status == 200:
                data = await response.json()
                return data.get('conversations', []), data.get('pages', {}).get('next', {}).get('starting_after')
            elif response.status == 429:  # Rate limit
                retry_after = int(response.headers.get('Retry-After', 60))
                await asyncio.sleep(retry_after)
                return await fetch_conversation_page(session, url, headers, payload, page_num)
            else:
                print(f"Error fetching page {page_num}: {response.status}")
                return [], None
    except Exception as e:
        print(f"Exception fetching page {page_num}: {e}")
        return [], None

async def search_conversations_async(
    start_date_str: str, 
    end_date_str: str, 
    product_area_filter_value: str | None = None,
    team_filter_details: dict | None = None
):
    """Asynchronous version of search_conversations."""
    if not INTERCOM_PROD_KEY:
        print("❌ FATAL: INTERCOM_PROD_KEY is not set. Cannot fetch conversations.")
        return None 

    start_timestamp = datetime.strptime(start_date_str, "%Y-%m-%d %H:%M").timestamp()
    end_timestamp = datetime.strptime(end_date_str, "%Y-%m-%d %H:%M").timestamp()

    url = "https://api.intercom.io/conversations/search"
    headers = {
        "Authorization": f"Bearer {INTERCOM_PROD_KEY}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    query_filters = [
        {"field": "statistics.last_close_at", "operator": ">", "value": int(start_timestamp)},
        {"field": "statistics.last_close_at", "operator": "<", "value": int(end_timestamp)}
    ]

    if product_area_filter_value:
        query_filters.append({
            "field": f"custom_attribute.{ACTUAL_PRODUCT_AREA_CUSTOM_ATTRIBUTE_NAME}",
            "operator": "=", 
            "value": product_area_filter_value
        })

    if team_filter_details:
        query_filters.append(team_filter_details)

    payload = {
        "query": {
            "operator": "AND",
            "value": query_filters
        },
        "pagination": {"per_page": 150}  # Increased from 100 to 150
    }

    all_conversations = []
    next_page = None
    page_num = 1

    async with aiohttp.ClientSession() as session:
        while True:
            conversations, next_page = await fetch_conversation_page(session, url, headers, payload, page_num)
            if conversations:
                all_conversations.extend(conversations)
                print(f"Page {page_num}: Fetched {len(conversations)} conversations. Total: {len(all_conversations)}")
            
            if not next_page:
                break
                
            payload['pagination']['starting_after'] = next_page
            page_num += 1

    return all_conversations

def search_conversations(
    start_date_str: str, 
    end_date_str: str, 
    product_area_filter_value: str | None = None,
    team_filter_details: dict | None = None
):
    """Synchronous wrapper for async search_conversations."""
    try:
        loop = asyncio.get_event_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
    return loop.run_until_complete(search_conversations_async(
        start_date_str, 
        end_date_str, 
        product_area_filter_value,
        team_filter_details
    ))

# ✅ Fetch full conversation details
async def get_intercom_conversation_async(session, conversation_id):
    """Asynchronous version of get_intercom_conversation."""
    url = f'https://api.intercom.io/conversations/{conversation_id}'
    retries = 3

    while retries > 0:
        try:
            async with session.get(url, headers={"Authorization": f"Bearer {INTERCOM_PROD_KEY}"}) as response:
                if response.status == 200:
                    return await response.json()
                elif response.status == 500:
                    print(f"⚠️ Server error. Retrying... ({retries} retries left)")
                    await asyncio.sleep(5)
                    retries -= 1
                else:
                    print(f"❌ Error fetching conversation {conversation_id}: {response.status}")
                    return None
        except Exception as e:
            print(f"❌ Request failed for conversation {conversation_id}: {e}")
            retries -= 1
            await asyncio.sleep(5)

    print(f"❌ Max retries reached for conversation {conversation_id}. Skipping.")
    return None

async def process_conversations_batch(conversations, batch_size=10):
    """Process conversations in batches asynchronously."""
    async with aiohttp.ClientSession() as session:
        for i in range(0, len(conversations), batch_size):
            batch = conversations[i:i + batch_size]
            tasks = [get_intercom_conversation_async(session, conv['id']) for conv in batch]
            results = await asyncio.gather(*tasks)
            yield [r for r in results if r is not None]

def filter_conversations_by_product(conversations, product):
    """Filter conversations by product area with improved performance."""
    filtered_conversations = []
    product_lower = product.lower()
    
    # Pre-filter conversations that don't have the required attribute
    valid_conversations = [
        conv for conv in conversations 
        if isinstance(conv.get('custom_attributes'), dict)
    ]
    
    # Process in batches
    for conv in valid_conversations:
        attributes = conv['custom_attributes']
        meta_mask_area = attributes.get('MetaMask area', '').strip()
        
        if meta_mask_area.lower() == product_lower:
            # Use cached version if available
            full_conversation = get_cached_conversation(conv['id'])
            if full_conversation:
                # Extract attributes more efficiently
                for category in CATEGORY_HEADERS.get(product, []):
                    full_conversation[category] = attributes.get(category, 'None')
                filtered_conversations.append(full_conversation)
    
    print(f"Total Conversations for {product}: {len(filtered_conversations)}")
    return filtered_conversations

# ✅ Store extracted data into an XLSX file
def store_conversations_to_xlsx(conversations, meta_mask_area, week_start_str_for_files, week_end_str_for_files):
    """Stores conversations in a dynamically named Excel file with improved performance."""
    if not conversations:
        print(f"No conversations to store for {meta_mask_area}.")
        return None

    file_name = f"{meta_mask_area.lower()}_conversations_{week_start_str_for_files}_to_{week_end_str_for_files}.xlsx"
    file_path = os.path.join(OUTPUT_DIR, file_name)

    # Pre-process data for better performance
    all_custom_attribute_keys = set()
    processed_rows = []
    
    for conv in conversations:
        if conv and isinstance(conv.get('custom_attributes'), dict):
            all_custom_attribute_keys.update(conv['custom_attributes'].keys())
    
    sorted_custom_attribute_keys = sorted(list(all_custom_attribute_keys))
    headers = ['conversation_id', 'summary', 'transcript'] + sorted_custom_attribute_keys
    
    # Process all rows at once
    for conversation in conversations:
        conversation_id = conversation.get('id', 'N/A')
        summary = sanitize_text(get_conversation_summary(conversation))
        transcript = sanitize_text(get_conversation_transcript(conversation))
        attributes = conversation.get('custom_attributes', {}) if isinstance(conversation.get('custom_attributes'), dict) else {}
        
        row_data = {
            'conversation_id': conversation_id,
            'summary': summary,
            'transcript': transcript
        }
        
        # Add all custom attributes efficiently
        for key in sorted_custom_attribute_keys:
            value = attributes.get(key, 'N/A')
            if isinstance(value, (list, dict)):
                value = str(value)
            row_data[key] = value
        
        processed_rows.append([row_data.get(header, 'N/A') for header in headers])

    # Create workbook and write data efficiently
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Conversations"
    
    # Write headers
    sheet.append(headers)
    
    # Write all rows at once
    for row in processed_rows:
        sheet.append(row)

    # Optimize column widths
    for col_idx, column_cells in enumerate(sheet.columns):
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap width at 100
        sheet.column_dimensions[column].width = adjusted_width

    # Apply text wrapping only to summary and transcript columns
    summary_col = openpyxl.utils.get_column_letter(headers.index('summary') + 1)
    transcript_col = openpyxl.utils.get_column_letter(headers.index('transcript') + 1)
    
    for col_letter in [summary_col, transcript_col]:
        for cell in sheet[col_letter]:
            cell.alignment = Alignment(wrap_text=True)

    workbook.save(file_path)
    print(f"✅ File saved: {file_path}")
    return file_path

# --- Helper function for Sentiment Analysis (moved to global scope) ---
def get_text_sentiment_tuple(text_input):
    if pd.isna(text_input) or not isinstance(text_input, str) or not text_input.strip():
        return 0.0, 0.0 
    try:
        blob = TextBlob(str(text_input)) 
        return blob.sentiment.polarity, blob.sentiment.subjectivity
    except Exception as e:
        # print(f"Warning: Could not process text for sentiment: '{str(text_input)[:50]}...'. Error: {e}")
        return 0.0, 0.0 

# ✅ Analyze XLSX and generate insights
def analyze_xlsx_and_generate_insights(xlsx_file, meta_mask_area, week_start_str_for_files, week_end_str_for_files):
    """Analyzes the Excel file with improved performance and accurate prompt responses."""
    try:
        if not xlsx_file or not os.path.exists(xlsx_file):
            print(f"Skipping analysis for {meta_mask_area}: XLSX file not found or not provided ({xlsx_file})")
            return None
        
        print(f"📊 Analyzing {xlsx_file} for {meta_mask_area}...")
        
        # Read Excel file with optimized settings
        df = pd.read_excel(
            xlsx_file,
            engine='openpyxl',
            dtype_backend='numpy_nullable'
        )
        df.columns = df.columns.str.strip()
        
        insights_file_name = f"{meta_mask_area.lower()}_insights_{week_start_str_for_files}_to_{week_end_str_for_files}.txt"
        insights_file_path = os.path.join(INSIGHTS_DIR, insights_file_name)
        
        if not os.path.exists(INSIGHTS_DIR):
            os.makedirs(INSIGHTS_DIR)
        
        analysis_text = [f"🚀 **Analysis for {meta_mask_area} ({week_start_str_for_files} to {week_end_str_for_files})**\n", "=" * 60, "\n\n"]
        
        # Pre-compute text source for better performance
        text_source_column = None
        if 'summary' in df.columns and not df['summary'].dropna().empty:
            valid_summaries = df['summary'][
                df['summary'].notna() & \
                ~df['summary'].str.strip().str.lower().isin(['', 'no summary available', 'n/a'])
            ]
            if not valid_summaries.empty:
                text_source_column = 'summary'
        
        if text_source_column is None and 'transcript' in df.columns and not df['transcript'].dropna().empty:
            valid_transcripts = df['transcript'][
                df['transcript'].notna() & \
                ~df['transcript'].str.strip().str.lower().isin(['', 'no transcript available', 'n/a'])
            ]
            if not valid_transcripts.empty:
                text_source_column = 'transcript'
        
        # Process all prompts efficiently
        prompts_to_process = {}
        global_prompts = PREDEFINED_PROMPTS.get("GLOBAL", {})
        area_specific_prompts = PREDEFINED_PROMPTS.get(meta_mask_area, {})
        
        for category, cat_prompts in global_prompts.items():
            if category not in prompts_to_process:
                prompts_to_process[category] = []
            prompts_to_process[category].extend(cat_prompts)
        
        for category, cat_prompts in area_specific_prompts.items():
            if category not in prompts_to_process:
                prompts_to_process[category] = []
            prompts_to_process[category].extend(cat_prompts)
        
        # Process each category
        for category_title, individual_prompts in prompts_to_process.items():
            try:
                analysis_text.append(f"\n🔹 **{category_title.replace('_', ' ')}**")
                
                for prompt_text in individual_prompts:
                    try:
                        analysis_text.append(f"\n  *Prompt: {prompt_text}*")
                        
                        # Process Keyword Analysis
                        if category_title == "Keyword Analysis":
                            if "What are the top 10 most important keywords?" in prompt_text:
                                if text_source_column:
                                    try:
                                        print(f"Processing keywords from {text_source_column} column...")
                                        text_for_keywords = df[text_source_column][df[text_source_column].notna()].astype(str)
                                        if not text_for_keywords.empty:
                                            all_text = " ".join(text_for_keywords.tolist()).lower()
                                            print(f"Processing {len(text_for_keywords)} text entries...")
                                            
                                            try:
                                                blob = TextBlob(all_text)
                                                print("TextBlob processing completed.")
                                                
                                                # Extract words with their POS tags and context
                                                word_contexts = []
                                                for sentence in blob.sentences:
                                                    try:
                                                        for word, tag in sentence.tags:
                                                            word_lower = word.lower()
                                                            if len(word_lower) > 2:  # Only consider words longer than 2 chars
                                                                # Calculate word importance based on:
                                                                # 1. Part of speech (nouns and adjectives are more important)
                                                                # 2. Word length (longer words tend to be more specific)
                                                                # 3. Frequency in the corpus
                                                                importance_score = 0
                                                                if tag.startswith('NN'):  # Nouns
                                                                    importance_score = 2
                                                                elif tag.startswith('JJ'):  # Adjectives
                                                                    importance_score = 1.5
                                                                elif tag.startswith('VB'):  # Verbs
                                                                    importance_score = 1
                                                                
                                                                # Adjust score based on word length
                                                                importance_score *= min(len(word_lower) / 4, 1.5)
                                                                
                                                                word_contexts.append({
                                                                    'word': word_lower,
                                                                    'score': importance_score,
                                                                    'pos': tag,
                                                                    'context': str(sentence)
                                                                })
                                                    except Exception as e:
                                                        print(f"Error processing sentence: {e}")
                                                        continue
                                                
                                                print(f"Processed {len(word_contexts)} word contexts.")
                                                
                                                # Convert to DataFrame for easier analysis
                                                word_df = pd.DataFrame(word_contexts)
                                                
                                                # Group by word and calculate aggregate scores
                                                word_scores = word_df.groupby('word').agg({
                                                    'score': 'sum',
                                                    'pos': lambda x: x.mode().iloc[0] if not x.empty else '',
                                                    'context': lambda x: x.iloc[0] if not x.empty else ''  # Keep first context
                                                }).reset_index()
                                                
                                                # Sort by score and get top 10
                                                top_words = word_scores.nlargest(10, 'score')
                                                
                                                analysis_text.append("\n  Top Keywords (with context and importance scores):")
                                                for _, row in top_words.iterrows():
                                                    word = row['word']
                                                    score = row['score']
                                                    pos = row['pos']
                                                    context = row['context']
                                                    # Truncate context to reasonable length
                                                    context = context[:100] + "..." if len(context) > 100 else context
                                                    analysis_text.append(f"    - {word} (Score: {score:.2f}, POS: {pos})")
                                                    analysis_text.append(f"      Context: {context}")
                                                
                                            except Exception as e:
                                                print(f"Error in TextBlob processing: {e}")
                                                analysis_text.append(f"    Error processing text: {str(e)}")
                                        else:
                                            analysis_text.append("    No valid text entries found for keyword analysis.")
                                    except Exception as e:
                                        print(f"Error processing text column: {e}")
                                        analysis_text.append(f"    Error processing text column: {str(e)}")
                                else:
                                    analysis_text.append("    No text source available for keyword extraction.")
                            
                            elif "What sentiment trends can be observed?" in prompt_text:
                                try:
                                    if text_source_column:
                                        # Group conversations by date for trend analysis
                                        df['date'] = pd.to_datetime(df['conversation_id'].str[:8], format='%Y%m%d', errors='coerce')
                                        
                                        # Calculate sentiment for each conversation
                                        sentiments = df[text_source_column].apply(get_text_sentiment_tuple)
                                        df['sentiment_polarity'] = sentiments.apply(lambda x: x[0] if isinstance(x, tuple) else 0.0)
                                        df['sentiment_subjectivity'] = sentiments.apply(lambda x: x[1] if isinstance(x, tuple) else 0.0)
                                        
                                        # Overall sentiment metrics
                                        avg_polarity = df['sentiment_polarity'].mean()
                                        avg_subjectivity = df['sentiment_subjectivity'].mean()
                                        
                                        analysis_text.append(f"\n  Overall Sentiment Analysis (from {text_source_column}):")
                                        analysis_text.append(f"    - Average Polarity: {avg_polarity:.2f} (range: -1 to 1)")
                                        analysis_text.append(f"    - Average Subjectivity: {avg_subjectivity:.2f} (range: 0 to 1)")
                                        
                                        # Sentiment distribution
                                        positive = len(df[df['sentiment_polarity'] > 0.1])
                                        neutral = len(df[(df['sentiment_polarity'] >= -0.1) & (df['sentiment_polarity'] <= 0.1)])
                                        negative = len(df[df['sentiment_polarity'] < -0.1])
                                        total = len(df)
                                        
                                        if total > 0:
                                            analysis_text.append("\n  Sentiment Distribution:")
                                            analysis_text.append(f"    - Positive: {positive} ({positive/total:.1%})")
                                            analysis_text.append(f"    - Neutral: {neutral} ({neutral/total:.1%})")
                                            analysis_text.append(f"    - Negative: {negative} ({negative/total:.1%})")
                                        
                                        # Sentiment trends over time
                                        if not df['date'].isna().all():
                                            daily_sentiment = df.groupby(df['date'].dt.date).agg({
                                                'sentiment_polarity': 'mean',
                                                'sentiment_subjectivity': 'mean',
                                                'conversation_id': 'count'
                                            }).reset_index()
                                            
                                            analysis_text.append("\n  Sentiment Trends Over Time:")
                                            for _, row in daily_sentiment.iterrows():
                                                date = row['date']
                                                polarity = row['sentiment_polarity']
                                                subjectivity = row['sentiment_subjectivity']
                                                count = row['conversation_id']
                                                analysis_text.append(f"    - {date}:")
                                                analysis_text.append(f"      Polarity: {polarity:.2f} (from {count} conversations)")
                                                analysis_text.append(f"      Subjectivity: {subjectivity:.2f}")
                                except Exception as e:
                                    print(f"Error in sentiment analysis: {e}")
                                    analysis_text.append(f"    Error in sentiment analysis: {str(e)}")
                    except Exception as e:
                        print(f"Error processing prompt '{prompt_text}': {e}")
                        analysis_text.append(f"    Error processing prompt: {str(e)}")
                        continue
            except Exception as e:
                print(f"Error processing category '{category_title}': {e}")
                analysis_text.append(f"Error processing category: {str(e)}")
                continue
        
        # Save insights file
        with open(insights_file_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(analysis_text))
        
        print(f"✅ Insights saved: {insights_file_path}")
        return insights_file_path
    
    except Exception as e:
        print(f"❌ Error in analyze_xlsx_and_generate_insights: {e}")
        return None

# Removed authenticate_google_drive and upload_to_google_drive functions

def _generate_scoped_product_area_files(
    conversations_for_scope: list, # List of 'thin' conversation objects for the current scope
    product_area_name: str,
    scope_identifier: str, # e.g., "TeamName" or "GLOBAL". Used for file naming.
    week_start_str_for_files: str,
    week_end_str_for_files: str,
    all_generated_files_list: list # List to append generated file paths to
) -> pd.DataFrame | None:
    """
    Filters conversations for a specific product area within a given scope,
    then generates and stores XLSX and insights files.

    Args:
        conversations_for_scope: List of conversation dicts (can be 'thin' from search).
        product_area_name: The name of the product area to filter for.
        scope_identifier: A string identifying the scope (e.g., team name or "GLOBAL").
        week_start_str_for_files: Date string for file naming.
        week_end_str_for_files: Date string for file naming.
        all_generated_files_list: List to which paths of generated files will be appended.

    Returns:
        A pandas DataFrame containing the data for the generated XLSX file, or None if no data.
    """
    print(f"  Filtering for product area '{product_area_name}' within scope '{scope_identifier}'...")
    # filter_conversations_by_product fetches full details if needed
    area_specific_conversations = filter_conversations_by_product(conversations_for_scope, product_area_name)

    if not area_specific_conversations:
        print(f"    No conversations found for product area '{product_area_name}' in scope '{scope_identifier}'.")
        return None

    file_label = f"{scope_identifier}_{product_area_name}" if scope_identifier != "GLOBAL" else product_area_name
    
    print(f"    Storing {len(area_specific_conversations)} conversations for '{file_label}'...")
    xlsx_file_path = store_conversations_to_xlsx(
        area_specific_conversations,
        file_label.replace(" ", "_"), # Ensure clean file name
        week_start_str_for_files,
        week_end_str_for_files
    )

    dataframe_for_report = None
    if xlsx_file_path:
        all_generated_files_list.append(xlsx_file_path)
        print(f"    Analyzing '{file_label}'...")
        insights_file_path = analyze_xlsx_and_generate_insights(
            xlsx_file_path,
            file_label.replace(" ", "_"), # Ensure clean context name
            week_start_str_for_files,
            week_end_str_for_files
        )
        if insights_file_path:
            all_generated_files_list.append(insights_file_path)
        
        try:
            dataframe_for_report = pd.read_excel(xlsx_file_path)
        except Exception as e:
            print(f"    Error reading {xlsx_file_path} back into DataFrame: {e}")
            dataframe_for_report = pd.DataFrame() # Return empty df on error
    
    return dataframe_for_report


def generate_end_of_shift_report(all_product_data, week_start_str_for_files, week_end_str_for_files):
    report_content = f"End of Shift Report ({week_start_str_for_files} to {week_end_str_for_files})\n\n"
    grand_total_conversations = 0
    all_issues_summary = {}

    for product_area, data in all_product_data.items():
        df = data['dataframe']
        if df is None or df.empty:
            report_content += f"No data processed for {product_area}.\n"
            continue

        num_conversations = len(df)
        grand_total_conversations += num_conversations
        report_content += f"Product Area: {product_area} ({num_conversations} conversations)\n"
        
        issue_columns = CATEGORY_HEADERS.get(product_area, [])
        primary_issue_col = issue_columns[0] if issue_columns else None

        if primary_issue_col and primary_issue_col in df.columns and not df[primary_issue_col].dropna().empty:
            top_issue = df[primary_issue_col].mode()[0] # Get the most frequent
            top_issue_count = df[primary_issue_col].value_counts().iloc[0]
            report_content += f"  - Biggest Issue: {top_issue} ({top_issue_count} occurrences)\n"
            
            # Aggregate for overall summary
            for issue, count in df[primary_issue_col].value_counts().items():
                all_issues_summary[f"{product_area} - {issue}"] = all_issues_summary.get(f"{product_area} - {issue}", 0) + count
        else:
            report_content += f"  - No primary issue data to report for {product_area}.\n"
        report_content += "---\n"

    report_content += f"\nOverall Summary:\n"
    report_content += f"Grand Total Conversations Processed: {grand_total_conversations}\n"
    
    if all_issues_summary:
        sorted_overall_issues = sorted(all_issues_summary.items(), key=lambda item: item[1], reverse=True)
        report_content += "Top 3 Issues Across All Products:\n"
        for i, (issue, count) in enumerate(sorted_overall_issues[:3]):
            report_content += f"  {i+1}. {issue}: {count} occurrences\n"
    else:
        report_content += "No specific issue data aggregated across products.\n"

    report_file_name = f"end_of_shift_report_{week_start_str_for_files}_to_{week_end_str_for_files}.txt"
    report_file_path = os.path.join(INSIGHTS_DIR, report_file_name)
    with open(report_file_path, 'w', encoding='utf-8') as f:
        f.write(report_content)
    print(f"✅ End of Shift Report saved: {report_file_path}")
    return report_file_path

def determine_conversation_team(conversation_data: dict) -> str:
    """Determines the team for a single conversation based on Intercom custom attributes."""
    attributes = conversation_data.get('custom_attributes', {})
    team_inbox_value = attributes.get(TEAM_INBOX_CUSTOM_ATTRIBUTE_KEY, "").strip()
    meta_mask_area = attributes.get('MetaMask area', '').strip().lower()

    # Priority 1: Direct mapping from Team Inbox value
    if team_inbox_value:
        team_inbox_lower = team_inbox_value.lower()
        # Direct matches for specific full Team Inbox values
        if team_inbox_lower == "metamask hd technical support" or team_inbox_lower == "metamask hd ts":
            return "MetaMask TS"
        if team_inbox_lower == "metamask hd ust":
            return "MetaMask HD UST"

        # Check for "MetaMask HD [TeamName]" pattern
        if team_inbox_lower.startswith("metamask hd "):
            potential_team_name = team_inbox_value.split(" ")[-1]
            if potential_team_name.lower() == "card":
                return "Card"
            elif potential_team_name.lower() == "portfolio":
                return "Portfolio"
            elif potential_team_name.lower() == "solana":
                return "Solana"

    # Priority 2: Fallback based on MetaMask Area
    if meta_mask_area == 'security':
        # Security area is handled by UST team
        return "MetaMask HD UST"
    
    # If Team Inbox had a value but didn't match any specific pattern above, log it
    if team_inbox_value:
        print(f"    ⚠️ Unmapped 'Team Inbox' value: '{team_inbox_value}'. Conversation may be misclassified or fall into General/Unclassified.")

    # Default/Catch-all
    return "MetaMask HD General"

def generate_team_end_of_shift_report(team_name, team_conversations, week_start_str_for_files, week_end_str_for_files):
    """Generates an End of Shift report specifically for a given team's conversations."""
    report_content = f"Team End of Shift Report: {team_name} ({week_start_str_for_files} to {week_end_str_for_files})\n"
    report_content += "=" * (len(report_content) -1) + "\n\n"

    if not team_conversations:
        report_content += "No conversations found for this team in this period.\n"
        # Save and return even if empty for consistency
        report_file_name = f"{team_name.replace(' ', '_')}_EOS_Report_{week_start_str_for_files}_to_{week_end_str_for_files}.txt"
        report_file_path = os.path.join(TEAM_REPORTS_DIR, report_file_name)
        with open(report_file_path, 'w', encoding='utf-8') as f:
            f.write(report_content)
        print(f"✅ Empty Team End of Shift Report saved: {report_file_path}")
        return report_file_path, report_content

    # Convert list of conversations to DataFrame for analysis similar to existing report
    # We need to extract relevant fields: custom attributes for issues, summaries.
    data_for_df = []
    for conv in team_conversations:
        attrs = conv.get('custom_attributes', {})
        summary = get_conversation_summary(conv) # Assuming this function is defined elsewhere
        row = {'summary': summary, **attrs} # Include all custom attributes
        data_for_df.append(row)
    df = pd.DataFrame(data_for_df)

    num_conversations = len(df)
    report_content += f"Total Conversations for Team: {num_conversations}\n\n"

    # Simplified: Aggregate top issues from all relevant CATEGORY_HEADERS for this team's conversations
    # This part needs careful thought: which columns define "issues" for a team?
    # For now, let's try to find most common values in *any* of the known issue columns present in this team's data.
    team_issues_summary = {}
    all_known_issue_columns = [col for cols in CATEGORY_HEADERS.values() for col in cols]
    
    for col_name in df.columns:
        if col_name in all_known_issue_columns and not df[col_name].dropna().empty:
            for issue, count in df[col_name].value_counts().items():
                if issue not in [None, "", "N/A", "None"]:
                    team_issues_summary[f"{col_name} - {issue}"] = team_issues_summary.get(f"{col_name} - {issue}", 0) + count
    
    if team_issues_summary:
        sorted_team_issues = sorted(team_issues_summary.items(), key=lambda item: item[1], reverse=True)
        report_content += "Top Issues/Custom Attributes Reported for Team:\n"
        for i, (issue, count) in enumerate(sorted_team_issues[:5]): # Top 5 overall for the team
            report_content += f"  {i+1}. {issue}: {count} occurrences\n"
    else:
        report_content += "No specific issue data aggregated for this team (check CATEGORY_HEADERS and data).\n"
    report_content += "---\n"
    
    # Top keywords from summaries for this team
    if 'summary' in df.columns and not df['summary'].dropna().empty:
        all_summaries = " ".join(df['summary'].dropna().astype(str).tolist()).lower()
        words = re.findall(r'\b\w+\b', all_summaries)
        filtered_words = [word for word in words if len(word) > 2]
        word_counts = pd.Series(filtered_words).value_counts().nlargest(5)
        
        report_content += "\nTop 5 Keywords from Summaries for Team:\n"
        for keyword, count in word_counts.items():
            report_content += f"- {keyword}: {count}\n"
    else:
        report_content += "\nNo summary data available for keyword analysis for this team.\n"
    
    team_text_source_column = None
    if 'summary' in df.columns and not df['summary'].dropna().empty:
        valid_summaries = df['summary'][df['summary'].notna() & ~df['summary'].str.strip().str.lower().isin(['', 'no summary available', 'n/a'])]
        if not valid_summaries.empty:
            team_text_source_column = 'summary'
    if team_text_source_column is None and 'transcript' in df.columns and not df['transcript'].dropna().empty:
        valid_transcripts = df['transcript'][df['transcript'].notna() & ~df['transcript'].str.strip().str.lower().isin(['', 'no transcript available', 'n/a'])]
        if not valid_transcripts.empty:
            team_text_source_column = 'transcript'

    if team_text_source_column:
        print(f"Calculating sentiment for team '{team_name}' from '{team_text_source_column}' column...")
        try:
            # Now calls the global get_text_sentiment_tuple
            team_sentiments = df[team_text_source_column].apply(get_text_sentiment_tuple) 
            df['sentiment_polarity'] = team_sentiments.apply(lambda x: x[0] if isinstance(x, tuple) else 0.0)
            df['sentiment_subjectivity'] = team_sentiments.apply(lambda x: x[1] if isinstance(x, tuple) else 0.0)

            avg_polarity = df['sentiment_polarity'].mean()
            avg_subjectivity = df['sentiment_subjectivity'].mean()
            
            report_content += f"\nSentiment Analysis (from {team_text_source_column}): \n"
            report_content += f"  - Average Polarity: {avg_polarity:.2f} \n"
            report_content += f"  - Average Subjectivity: {avg_subjectivity:.2f} \n"

            positive_threshold = 0.05
            negative_threshold = -0.05
            positive_count = df[df['sentiment_polarity'] > positive_threshold].shape[0]
            negative_count = df[df['sentiment_polarity'] < negative_threshold].shape[0]
            neutral_count = df[(df['sentiment_polarity'] >= negative_threshold) & (df['sentiment_polarity'] <= positive_threshold)].shape[0]
            total_sentiments = positive_count + negative_count + neutral_count

            if total_sentiments > 0:
                report_content += "  Distribution: \n"
                report_content += f"    - Positive: {positive_count} ({positive_count/total_sentiments:.1%}) \n"
                report_content += f"    - Neutral:  {neutral_count} ({neutral_count/total_sentiments:.1%}) \n"
                report_content += f"    - Negative: {negative_count} ({negative_count/total_sentiments:.1%}) \n"
            else:
                report_content += "  No sentiment scores available for distribution for this team. \n"
        except Exception as e:
            report_content += f"\nError during team sentiment analysis: {e} \n"
            print(f"Error in team sentiment for {team_name}: {e}")
    else:
        report_content += "\nSentiment analysis not performed for team (no suitable text source found). \n"
    report_content += "---\n"
    
    # ... (rest of team report generation, including keyword analysis if any, and saving the file) ...

    report_file_name = f"{team_name.replace(' ', '_')}_EOS_Report_{week_start_str_for_files}_to_{week_end_str_for_files}.txt"
    report_file_path = os.path.join(TEAM_REPORTS_DIR, report_file_name)
    with open(report_file_path, 'w', encoding='utf-8') as f:
        f.write(report_content)
    print(f"✅ Team End of Shift Report saved: {report_file_path}")
    return report_file_path, report_content

def get_date_range(start_date: str | None = None, end_date: str | None = None) -> tuple[str, str]:
    """
    Get the date range for processing. If no dates provided, defaults to last week.
    
    Args:
        start_date: Optional start date in format "YYYY-MM-DD HH:MM"
        end_date: Optional end date in format "YYYY-MM-DD HH:MM"
    
    Returns:
        Tuple of (start_date, end_date) in format "YYYY-MM-DD HH:MM"
    """
    if start_date and end_date:
        return start_date, end_date
        
    # Default to last week if no dates provided
    today = datetime.now()
    last_monday = today - timedelta(days=today.weekday() + 7)
    last_sunday = last_monday + timedelta(days=6)
    
    return (
        last_monday.strftime("%Y-%m-%d 00:00"),
        last_sunday.strftime("%Y-%m-%d 23:59")
    )

async def main_function(
    start_date_str: str, 
    end_date_str: str, 
    send_to_slack: bool = False,
    target_team_name: str | None = None,
    target_product_area_name: str | None = None
):
    """Main function to fetch, process, and analyze Intercom conversations, with targeted reporting."""
    logger.info("=== Starting LLM5.py main_function ===")
    logger.info(f"Parameters: start_date={start_date_str}, end_date={end_date_str}, slack={send_to_slack}")
    logger.info(f"Target team: {target_team_name}, Target product area: {target_product_area_name}")

    # Initialize tracking variables
    all_generated_files = []
    processed_counts = {
        "total_conversations_fetched": 0,
        "targeted_team_product_area_files": 0,
        "team_specific_product_area_files": 0,
        "global_product_area_files": 0,
        "team_eos_reports_generated": 0
    }

    try:
        # --- Initial Critical Checks ---
        logger.info("=== Checking Environment Variables ===")
        if not INTERCOM_PROD_KEY:
            error_msg = "CRITICAL ERROR: INTERCOM_PROD_KEY is not set in environment variables."
            logger.error(error_msg)
            return {"status": "failed", "message": error_msg, "local_files": [], "processed_counts": processed_counts}

        logger.info("✅ INTERCOM_PROD_KEY is set")

        # --- Date Validation ---
        logger.info("=== Validating Dates ===")
        try:
            start_date_dt = datetime.strptime(start_date_str, "%Y-%m-%d %H:%M")
            end_date_dt = datetime.strptime(end_date_str, "%Y-%m-%d %H:%M")
            logger.debug(f"Parsed dates - Start: {start_date_dt}, End: {end_date_dt}")
            
            if start_date_dt > end_date_dt:
                error_msg = f"Input Error: Start date ({start_date_str}) must be before or same as end date ({end_date_str})."
                logger.error(error_msg)
                return {"status": "failed", "message": error_msg, "local_files": [], "processed_counts": processed_counts}
            logger.info("✅ Date validation passed")
        except ValueError as e:
            error_msg = f"Date format error: {str(e)}"
            logger.error(error_msg)
            return {"status": "failed", "message": error_msg, "local_files": [], "processed_counts": processed_counts}

        # --- File System Checks ---
        logger.info("=== Checking File System ===")
        try:
            for directory in [OUTPUT_DIR, INSIGHTS_DIR, TEAM_REPORTS_DIR]:
                if not os.path.exists(directory):
                    logger.info(f"Creating directory: {directory}")
                    os.makedirs(directory)
                logger.debug(f"Directory exists: {directory}")
        except Exception as e:
            error_msg = f"File system error: {str(e)}"
            logger.error(error_msg)
            return {"status": "failed", "message": error_msg, "local_files": [], "processed_counts": processed_counts}

        # --- API Filter Setup ---
        logger.info("=== Setting Up API Filters ===")
        try:
            product_area_api_filter_value = None
            team_api_filter_details = None

            if target_product_area_name and target_product_area_name != "ALL_AREAS":
                product_area_api_filter_value = target_product_area_name
                logger.info(f"Product Area filter: {product_area_api_filter_value}")

            if target_team_name:
                logger.info(f"Setting up team filter for: {target_team_name}")
                if target_team_name in SCRIPT_TEAM_TO_TEAM_INBOX_VALUE:
                    team_inbox_filter_value = SCRIPT_TEAM_TO_TEAM_INBOX_VALUE[target_team_name]
                    team_api_filter_details = {
                        "field": f"custom_attribute.{TEAM_INBOX_CUSTOM_ATTRIBUTE_KEY}",
                        "operator": "=", 
                        "value": team_inbox_filter_value
                    }
                    logger.info(f"Team filter set using Team Inbox value: {team_inbox_filter_value}")
                else:
                    logger.warning(f"Team not found in SCRIPT_TEAM_TO_TEAM_INBOX_VALUE, will try fallback methods")
        except Exception as e:
            error_msg = f"Error setting up API filters: {str(e)}"
            logger.error(error_msg)
            return {"status": "failed", "message": error_msg, "local_files": [], "processed_counts": processed_counts}

        # --- Fetch Conversations ---
        logger.info("=== Fetching Conversations ===")
        try:
            logger.info("Starting conversation search...")
            conversations = search_conversations(
                start_date_str, 
                end_date_str, 
                product_area_filter_value=product_area_api_filter_value,
                team_filter_details=team_api_filter_details
            )
            
            if conversations is None:
                error_msg = "Failed to fetch Intercom conversations after retries. Please check logs and API key."
                logger.error(error_msg)
                return {"status": "failed", "message": error_msg, "local_files": all_generated_files, "processed_counts": processed_counts}

            if not conversations:
                error_msg = "No conversations found for the selected timeframe and filters."
                logger.warning(error_msg)
                processed_counts["total_conversations_fetched"] = 0 
                return {"status": "no_data", "message": error_msg, "local_files": all_generated_files, "processed_counts": processed_counts}

            processed_counts["total_conversations_fetched"] = len(conversations)
            logger.info(f"Successfully fetched {len(conversations)} conversations")
        except Exception as e:
            error_msg = f"Error fetching conversations: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {"status": "failed", "message": error_msg, "local_files": all_generated_files, "processed_counts": processed_counts}

        # --- Process Conversations ---
        logger.info("=== Processing Conversations ===")
        try:
            logger.info("Grouping conversations by team...")
            team_grouped_conversations = {
                "MetaMask TS": [], "MetaMask HD UST": [], "Card": [], 
                "Portfolio": [], "Solana": [], "MetaMask HD General": [],
                "Unclassified": [] 
            }
            
            for conv_data in conversations:
                if isinstance(conv_data, dict):
                    team = determine_conversation_team(conv_data)
                    team_grouped_conversations.get(team, team_grouped_conversations["Unclassified"]).append(conv_data)
                else:
                    print(f"⚠️ Warning: Found non-dictionary item in conversations list: {type(conv_data)}")

            logger.info("Team grouping complete:")
            for team, convs in team_grouped_conversations.items():
                logger.info(f"  - {team}: {len(convs)} conversations")
        except Exception as e:
            error_msg = f"Error processing conversations: {str(e)}"
            logger.error(error_msg)
            return {"status": "failed", "message": error_msg, "local_files": all_generated_files, "processed_counts": processed_counts}

        # --- Generate Reports ---
        logger.info("=== Generating Reports ===")
        try:
            if target_team_name and target_product_area_name and target_product_area_name != "ALL_AREAS":
                logger.info(f"Generating targeted report for team '{target_team_name}' and product area '{target_product_area_name}'")
                team_convs = team_grouped_conversations.get(target_team_name, [])
                df = _generate_scoped_product_area_files(
                    team_convs,
                    target_product_area_name,
                    target_team_name,
                    start_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                    end_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                    all_generated_files
                )
                processed_counts["targeted_team_product_area_files"] += 1 if df is not None else 0
            elif target_team_name:
                logger.info(f"Generating team-specific reports for '{target_team_name}'")
                team_convs = team_grouped_conversations.get(target_team_name, [])
                for product_area in CATEGORY_HEADERS.keys():
                    df = _generate_scoped_product_area_files(
                        team_convs,
                        product_area,
                        target_team_name,
                        start_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                        end_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                        all_generated_files
                    )
                    processed_counts["team_specific_product_area_files"] += 1 if df is not None else 0
                # Team End of Shift report
                team_report_path, _ = generate_team_end_of_shift_report(
                    target_team_name, team_convs,
                    start_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                    end_date_str.replace("-", "").replace(" ", "_").replace(":", "")
                )
                all_generated_files.append(team_report_path)
                processed_counts["team_eos_reports_generated"] += 1
            elif target_product_area_name and target_product_area_name != "ALL_AREAS":
                logger.info(f"Generating product area report for '{target_product_area_name}'")
                df = _generate_scoped_product_area_files(
                    conversations,
                    target_product_area_name,
                    "GLOBAL",
                    start_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                    end_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                    all_generated_files
                )
                processed_counts["global_product_area_files"] += 1 if df is not None else 0
            else:
                logger.info("Generating full global reports")
                all_product_data = {}
                for product_area in CATEGORY_HEADERS.keys():
                    df = _generate_scoped_product_area_files(
                        conversations,
                        product_area,
                        "GLOBAL",
                        start_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                        end_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                        all_generated_files
                    )
                    all_product_data[product_area] = {"dataframe": df}
                    processed_counts["global_product_area_files"] += 1 if df is not None else 0
                # End of Shift report
                eos_report_path = generate_end_of_shift_report(
                    all_product_data,
                    start_date_str.replace("-", "").replace(" ", "_").replace(":", ""),
                    end_date_str.replace("-", "").replace(" ", "_").replace(":", "")
                )
                all_generated_files.append(eos_report_path)
        except Exception as e:
            error_msg = f"Error generating reports: {str(e)}"
            logger.error(error_msg)
            return {"status": "failed", "message": error_msg, "local_files": all_generated_files, "processed_counts": processed_counts}

        logger.info("=== Processing Complete ===")
        return {
            "status": "success",
            "message": "Processing completed successfully.",
            "local_files": all_generated_files,
            "processed_counts": processed_counts
        }

    except Exception as e:
        error_msg = f"An unexpected error occurred: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {
            "status": "failed",
            "message": error_msg,
            "local_files": all_generated_files,
            "processed_counts": processed_counts
        }

async def process_single_conversation(conversation_id: str) -> None:
    """
    Process a single conversation asynchronously.
    
    Args:
        conversation_id: The Intercom conversation ID to process
    """
    try:
        logger.info(f"Processing single conversation: {conversation_id}")
        async with aiohttp.ClientSession() as session:
            conversation_data = await get_intercom_conversation_async(session, conversation_id)
            
            if not conversation_data:
                error_msg = f"❌ Could not fetch conversation ID: {conversation_id}"
                logger.error(error_msg)
                print(error_msg)
                return
            
            attributes = conversation_data.get('custom_attributes', {})
            meta_mask_area_single = attributes.get('MetaMask area', 'Unknown_Area').strip()
            if not meta_mask_area_single: 
                meta_mask_area_single = "Unknown_Area"
            
            logger.info(f"Conversation belongs to MetaMask Area: {meta_mask_area_single}")
            file_timestamp_str = conversation_id

            logger.info(f"Storing conversation {conversation_id} for area {meta_mask_area_single}...")
            xlsx_file_path_single = store_conversations_to_xlsx(
                [conversation_data], meta_mask_area_single, 
                file_timestamp_str, "single"
            )
            
            all_generated_files_single = []
            
            if xlsx_file_path_single:
                all_generated_files_single.append(xlsx_file_path_single)
                logger.info(f"Analyzing {xlsx_file_path_single}...")
                insights_file_path_single = analyze_xlsx_and_generate_insights(
                    xlsx_file_path_single, meta_mask_area_single, 
                    file_timestamp_str, "single"
                )
                if insights_file_path_single:
                    all_generated_files_single.append(insights_file_path_single)
            
            logger.info("\n--- Single Conversation Test Run Summary ---")
            logger.info(f"Processed Conversation ID: {conversation_id}")
            logger.info("Local files generated:")
            for f in all_generated_files_single:
                logger.info(f"  - {f}")
    except Exception as e:
        error_msg = f"Error processing conversation {conversation_id}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        print(error_msg)

if __name__ == "__main__":
    try:
        # Set up console output to show all logs
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
        
        logger.info("=== Starting LLM5.py script ===")
        logger.info(f"Python version: {sys.version}")
        logger.info(f"Current working directory: {os.getcwd()}")
        logger.info(f"Script location: {os.path.abspath(__file__)}")
        
        # Verify required directories exist
        for directory in [OUTPUT_DIR, INSIGHTS_DIR, TEAM_REPORTS_DIR]:
            if not os.path.exists(directory):
                logger.info(f"Creating directory: {directory}")
                os.makedirs(directory)
        
        parser = argparse.ArgumentParser(description="Fetch and analyze Intercom conversations.")
        parser.add_argument("-c", "--conversation_id", type=str, help="Fetch a single conversation by ID (generates its area files).")
        parser.add_argument("--start_date", type=str, help="Start date (YYYY-MM-DD HH:MM).")
        parser.add_argument("--end_date", type=str, help="End date (YYYY-MM-DD HH:MM).")
        parser.add_argument("--send_slack", action="store_true", help="Send generated team reports to Slack.")
        parser.add_argument("--target-team", type=str, help="Specify a single team name to generate reports for.")
        parser.add_argument("--target-product-area", type=str, help="Specify a single product area to generate reports for.")
        parser.add_argument("--suggest-stop-words", action="store_true", help="Analyze conversation data from a previous run to suggest stop words.")
        parser.add_argument("--stop-words-input-dir", type=str, default=OUTPUT_DIR, help="Directory containing XLSX files to scan for stop word suggestion.")

        args = parser.parse_args()
        logger.info(f"Parsed arguments: {vars(args)}")

        # Verify environment variables
        if not INTERCOM_PROD_KEY:
            error_msg = "🛑 INTERCOM_PROD_KEY not found in .env. Please set it."
            logger.error(error_msg)
            print(error_msg)
            sys.exit(1)

        if args.conversation_id:
            logger.info(f"Fetching single conversation ID: {args.conversation_id}")
            try:
                asyncio.run(process_single_conversation(args.conversation_id))
            except Exception as e:
                error_msg = f"Error processing conversation: {str(e)}"
                logger.error(error_msg, exc_info=True)
                print(error_msg)
                sys.exit(1)
        else:
            # Get date range using the new function
            try:
                run_start_date, run_end_date = get_date_range(args.start_date, args.end_date)
                logger.info(f"Date range calculated: {run_start_date} to {run_end_date}")
            except Exception as e:
                error_msg = f"Error calculating date range: {str(e)}"
                logger.error(error_msg, exc_info=True)
                print(error_msg)
                sys.exit(1)
            
            logger.info(f"Running LLM5.py for date range: {run_start_date} to {run_end_date}")
            if args.target_team: logger.info(f"Targeting Team: {args.target_team}")
            if args.target_product_area: logger.info(f"Targeting Product Area: {args.target_product_area}")

            try:
                result = asyncio.run(main_function(
                    run_start_date, 
                    run_end_date, 
                    send_to_slack=args.send_slack,
                    target_team_name=args.target_team,
                    target_product_area_name=args.target_product_area
                ))
                
                logger.info("\n--- Batch Run Summary ---")
                logger.info(f"Status: {result.get('status')}")
                logger.info(f"Message: {result.get('message')}")
                logger.info("Local files generated:")
                for f in result.get('local_files', []): logger.info(f"  - {f}")
                logger.info("Processed Counts:")
                for k, v in result.get('processed_counts', {}).items(): logger.info(f"  - {k}: {v}")
                
                if result.get('status') == 'failed':
                    error_msg = f"Script failed: {result.get('message')}"
                    logger.error(error_msg)
                    print(error_msg)
                    sys.exit(1)
                    
            except Exception as e:
                error_msg = f"Error in main_function: {str(e)}"
                logger.error(error_msg, exc_info=True)
                print(error_msg)
                sys.exit(1)
                
    except Exception as e:
        error_msg = f"Unexpected error in script: {str(e)}"
        logger.error(error_msg, exc_info=True)
        print(error_msg)
        sys.exit(1)

# Add missing function
def get_intercom_conversation(conversation_id):
    """Fetch a single conversation from Intercom API."""
    url = f'https://api.intercom.io/conversations/{conversation_id}'
    headers = {"Authorization": f"Bearer {INTERCOM_PROD_KEY}"}
    
    try:
        response = SESSION.get(url, headers=headers)
        if response.status_code == 200:
            return response.json()
        elif response.status_code == 429:  # Rate limit
            retry_after = int(response.headers.get('Retry-After', 60))
            time.sleep(retry_after)
            return get_intercom_conversation(conversation_id)  # Retry
        else:
            print(f"Error fetching conversation {conversation_id}: {response.status_code}")
            return None
    except Exception as e:
        print(f"Request failed for conversation {conversation_id}: {e}")
        return None
